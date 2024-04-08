import openpyxl
import json
import os
    
try:
    excelPath = "." #使用ディレクトリによって書き換える
    jsonPath = "." #使用ディレクトリによって書き換える
    files = os.listdir(excelPath)
    for excelfile in files:
        if excelfile.endswith('.xlsx'):
            book = openpyxl.load_workbook(excelPath + '/' + excelfile)

            #シートの枚数分ループ処理を行う
            for sheetIndex in range(len(book.sheetnames)):
                sheet = book.worksheets[sheetIndex] #シート
                columnCount = sheet.max_column      #列数
                rowCount = sheet.max_row            #行数
                columnNames = []                    #カラム名配列
                dataList = []                       #データを格納する配列

                #カラム名の抽出
                for columnIndex in range(1, columnCount+1):
                    columnNames += [sheet.cell(column=columnIndex, row=2).value]

                #データをカラム名と紐づけて連想配列に格納
                for rowIndex in range(3, rowCount+1):
                    valueDict = {}
                    for columnIndex in range(1, columnCount+1):
                        cellValue = sheet.cell(column=columnIndex, row=rowIndex).value
                        valueDict[columnNames[columnIndex-1]] = cellValue
                    dataList += [valueDict]
                
                #JSONファイルに出力
                fileTitle = sheet.title + '.json'
                # jsonData = {sheet.title + 'Model':dataList}
                with open(jsonPath + fileTitle, 'w', encoding='utf_8_sig') as f:
                    json.dump(dataList, f, ensure_ascii=False, indent=4)
                print(fileTitle)
    print('success')

except Exception as e:
    print(e)